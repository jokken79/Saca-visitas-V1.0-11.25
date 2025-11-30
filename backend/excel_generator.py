# ============================================================
# UNS VISA SYSTEM - Excel Generator
# Genera申請書 en formato oficial de出入国在留管理庁
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from typing import Optional, Dict, Any
import io

class VisaFormExcelGenerator:
    """
    Generador de formularios de visa en formato Excel oficial de入管
    
    Soporta:
    - 在留期間更新許可申請書 (Renewal)
    - 在留資格認定証明書交付申請書 (COE)
    - 在留資格変更許可申請書 (Change)
    """
    
    def __init__(self):
        # Estilos estándar
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        self.highlight_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        self.title_font = Font(name='MS Gothic', size=14, bold=True)
        self.header_font = Font(name='MS Gothic', size=10, bold=True)
        self.normal_font = Font(name='MS Gothic', size=9)
        self.small_font = Font(name='MS Gothic', size=8)
        
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.top_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def generate_renewal_form(self, data: Dict[str, Any]) -> io.BytesIO:
        """
        Genera 在留期間更新許可申請書 (技術・人文知識・国際業務)
        
        Args:
            data: Diccionario con todos los datos del formulario
            
        Returns:
            BytesIO con el archivo Excel
        """
        wb = Workbook()
        
        # Crear las 4 hojas del formulario oficial
        self._create_applicant_sheet_1(wb, data)
        self._create_applicant_sheet_2(wb, data)
        self._create_applicant_sheet_3(wb, data)
        self._create_organization_sheet(wb, data)
        
        # Eliminar la hoja por defecto
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Guardar en BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_applicant_sheet_1(self, wb: Workbook, data: Dict):
        """申請人等作成用１ - Información básica"""
        ws = wb.create_sheet('申請人等作成用１')
        
        # Configurar anchos de columna (simular formato A4)
        column_widths = [4, 12, 12, 12, 12, 10, 10, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # ===== ENCABEZADO =====
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = '在 留 期 間 更 新 許 可 申 請 書'
        cell.font = self.title_font
        cell.alignment = self.center_align
        
        ws.merge_cells('A2:H2')
        ws['A2'].value = 'APPLICATION FOR EXTENSION OF PERIOD OF STAY'
        ws['A2'].font = Font(name='Arial', size=9)
        ws['A2'].alignment = self.center_align
        
        # Oficina de presentación
        ws.merge_cells('A4:D4')
        office = data.get('submission_office', '名古屋')
        ws['A4'].value = f'{office} 出入国在留管理局長 殿'
        ws['A4'].font = self.normal_font
        
        # Área de foto
        ws.merge_cells('G3:H8')
        photo_cell = ws['G3']
        photo_cell.value = '写真\nPhoto\n縦4cm×横3cm\n\n(申請前3月以内に\n撮影したもの)'
        photo_cell.font = self.small_font
        photo_cell.alignment = self.center_align
        photo_cell.border = self.thin_border
        
        # ===== SECCIÓN: 申請人等作成用１ =====
        ws.merge_cells('A6:F6')
        ws['A6'].value = '【申請人等作成用１】 FOR APPLICANT, PART 1'
        ws['A6'].font = self.header_font
        ws['A6'].fill = self.header_fill
        
        row = 9
        
        # 1. Nacionalidad
        self._add_form_row(ws, row, '1', '国籍・地域', 'Nationality/Region', 
                         data.get('nationality', ''))
        row += 1
        
        # 2. Fecha de nacimiento
        dob = data.get('date_of_birth', '')
        if isinstance(dob, date):
            dob = dob.strftime('%Y年%m月%d日')
        self._add_form_row(ws, row, '2', '生年月日', 'Date of birth', dob)
        row += 1
        
        # 3. Nombre
        full_name = f"{data.get('family_name', '')} {data.get('given_name', '')}"
        kanji = data.get('name_kanji', '')
        if kanji:
            full_name = f"{kanji} ({full_name})"
        self._add_form_row(ws, row, '3', '氏名', 'Name', full_name)
        row += 1
        
        # 4. Sexo
        sex = data.get('sex', '')
        sex_display = '☑ 男 Male  ☐ 女 Female' if sex == 'male' else '☐ 男 Male  ☑ 女 Female'
        self._add_form_row(ws, row, '4', '性別', 'Sex', sex_display)
        row += 1
        
        # 5. Estado civil
        marital = data.get('marital_status', '')
        marital_display = '☑ 有 Married  ☐ 無 Single' if marital == 'married' else '☐ 有 Married  ☑ 無 Single'
        self._add_form_row(ws, row, '5', '配偶者の有無', 'Marital status', marital_display)
        row += 1
        
        # 6. Ocupación
        self._add_form_row(ws, row, '6', '職業', 'Occupation', 
                         data.get('occupation', '会社員'))
        row += 1
        
        # 7. Ciudad natal
        self._add_form_row(ws, row, '7', '本国における居住地', 'Home town/city', 
                         data.get('home_town_city', ''))
        row += 1
        
        # 8. Dirección en Japón
        self._add_form_row_wide(ws, row, '8', '住居地', 'Address in Japan', 
                               data.get('address_japan', ''))
        row += 1
        
        # 9. Teléfono
        ws[f'A{row}'].value = '9'
        ws[f'A{row}'].font = self.small_font
        ws[f'B{row}'].value = '電話番号'
        ws[f'C{row}'].value = data.get('telephone_japan', '')
        ws[f'D{row}'].value = '携帯電話番号'
        ws[f'E{row}'].value = data.get('cellular_phone', '')
        row += 1
        
        # 10. Pasaporte
        ws[f'A{row}'].value = '10'
        ws[f'B{row}'].value = '旅券'
        ws[f'C{row}'].value = f"番号: {data.get('passport_number', '')}"
        exp = data.get('passport_expiration', '')
        if isinstance(exp, date):
            exp = exp.strftime('%Y年%m月%d日')
        ws[f'E{row}'].value = f"有効期限: {exp}"
        row += 1
        
        # 11. Estado de residencia actual
        ws[f'A{row}'].value = '11'
        ws[f'B{row}'].value = '現に有する在留資格'
        ws[f'C{row}'].value = data.get('current_visa_status', '技術・人文知識・国際業務')
        ws[f'E{row}'].value = f"在留期間: {data.get('current_period_of_stay', '')}"
        exp_date = data.get('current_expiration_date', '')
        if isinstance(exp_date, date):
            exp_date = exp_date.strftime('%Y年%m月%d日')
        ws[f'G{row}'].value = f"満了日: {exp_date}"
        row += 1
        
        # 12. Número de tarjeta de residencia
        self._add_form_row(ws, row, '12', '在留カード番号', 'Residence card number', 
                         data.get('residence_card_number', ''))
        row += 1
        
        # 13. Estado deseado
        ws[f'A{row}'].value = '13'
        ws[f'B{row}'].value = '希望する在留資格'
        ws[f'C{row}'].value = data.get('desired_visa_status', '技術・人文知識・国際業務')
        ws[f'E{row}'].value = f"在留期間: {data.get('desired_period', '5年')}"
        row += 1
        
        # 14. Razón de renovación
        self._add_form_row_wide(ws, row, '14', '更新の理由', 'Reason for extension', 
                               data.get('reason_for_extension', '引き続き日本で就労するため'))
        row += 1
        
        # 15. Antecedentes penales
        criminal = data.get('has_criminal_record', False)
        criminal_display = '☑ 有 Yes  ☐ 無 No' if criminal else '☐ 有 Yes  ☑ 無 No'
        self._add_form_row_wide(ws, row, '15', '犯罪を理由とする処分を受けたことの有無', 
                               'Criminal record', criminal_display)
        
        # Aplicar estilos por defecto
        for r in range(1, row + 2):
            for c in range(1, 9):
                cell = ws.cell(row=r, column=c)
                if not cell.font:
                    cell.font = self.normal_font
    
    def _create_applicant_sheet_2(self, wb: Workbook, data: Dict):
        """申請人等作成用２ - Familia y educación"""
        ws = wb.create_sheet('申請人等作成用２')
        
        # Configurar anchos
        for i, width in enumerate([4, 10, 10, 10, 8, 8, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Encabezado
        ws.merge_cells('A1:H1')
        ws['A1'].value = '【申請人等作成用２】 FOR APPLICANT, PART 2'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        
        row = 3
        
        # 16. Familia en Japón
        ws[f'A{row}'].value = '16'
        ws[f'B{row}'].value = '在日親族（父・母・配偶者・子・兄弟姉妹等）及び同居者'
        ws[f'B{row}'].font = self.normal_font
        row += 1
        
        has_family = data.get('has_family_in_japan', False)
        ws[f'B{row}'].value = '☑ 有  ☐ 無' if has_family else '☐ 有  ☑ 無'
        row += 1
        
        if has_family and data.get('family_in_japan'):
            # Encabezados de tabla
            headers = ['続柄', '氏名', '生年月日', '国籍', '同居', '勤務先', '在留カード番号']
            for i, h in enumerate(headers, 2):
                ws.cell(row=row, column=i).value = h
                ws.cell(row=row, column=i).font = self.small_font
                ws.cell(row=row, column=i).border = self.thin_border
            row += 1
            
            for member in data['family_in_japan']:
                ws.cell(row=row, column=2).value = member.get('relationship', '')
                ws.cell(row=row, column=3).value = member.get('name', '')
                ws.cell(row=row, column=4).value = member.get('date_of_birth', '')
                ws.cell(row=row, column=5).value = member.get('nationality', '')
                ws.cell(row=row, column=6).value = '予定' if member.get('residing_with') else ''
                ws.cell(row=row, column=7).value = member.get('place_of_employment', '')
                ws.cell(row=row, column=8).value = member.get('residence_card_number', '')
                row += 1
        
        row += 1
        
        # 17. Lugar de trabajo
        ws[f'A{row}'].value = '17'
        ws[f'B{row}'].value = '勤務先'
        ws[f'B{row}'].font = self.header_font
        row += 1
        
        ws[f'B{row}'].value = '(1) 名称'
        ws.merge_cells(f'C{row}:F{row}')
        ws[f'C{row}'].value = data.get('company_name', '')
        row += 1
        
        ws[f'B{row}'].value = '(2) 所在地'
        ws.merge_cells(f'C{row}:H{row}')
        ws[f'C{row}'].value = data.get('company_address', '')
        row += 1
        
        ws[f'B{row}'].value = '(3) 電話番号'
        ws[f'C{row}'].value = data.get('company_telephone', '')
        row += 2
        
        # 18. Educación
        ws[f'A{row}'].value = '18'
        ws[f'B{row}'].value = '最終学歴'
        ws[f'B{row}'].font = self.header_font
        row += 1
        
        school_loc = data.get('school_location', 'overseas')
        ws[f'B{row}'].value = '(1) 学校の所在地'
        ws[f'C{row}'].value = '☑ 日本  ☐ 外国' if school_loc == 'japan' else '☐ 日本  ☑ 外国'
        row += 1
        
        ws[f'B{row}'].value = '(2) 学校の種類'
        school_types = {
            'university_grad': '大学院', 'university': '大学',
            'junior_college': '短期大学', 'vocational': '専門学校'
        }
        ws[f'C{row}'].value = school_types.get(data.get('school_type', ''), '')
        row += 1
        
        ws[f'B{row}'].value = '(3) 学校名'
        ws.merge_cells(f'C{row}:F{row}')
        ws[f'C{row}'].value = data.get('school_name', '')
        row += 1
        
        ws[f'B{row}'].value = '(4) 卒業年月日'
        grad = data.get('graduation_date', '')
        if isinstance(grad, date):
            grad = grad.strftime('%Y年%m月%d日')
        ws[f'C{row}'].value = grad
        row += 1
        
        # 19. Campo de especialización
        ws[f'A{row}'].value = '19'
        ws[f'B{row}'].value = '専攻・専門分野'
        ws[f'C{row}'].value = data.get('major_field', '')
        row += 1
        
        # 20. Calificación IT
        ws[f'A{row}'].value = '20'
        ws[f'B{row}'].value = '情報処理資格'
        has_it = data.get('has_it_qualification', False)
        if has_it:
            ws[f'C{row}'].value = f"☑ 有 ({data.get('it_qualification_name', '')})  ☐ 無"
        else:
            ws[f'C{row}'].value = '☐ 有  ☑ 無'
    
    def _create_applicant_sheet_3(self, wb: Workbook, data: Dict):
        """申請人等作成用３ - Historial laboral"""
        ws = wb.create_sheet('申請人等作成用３')
        
        for i, width in enumerate([4, 20, 15, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.merge_cells('A1:D1')
        ws['A1'].value = '【申請人等作成用３】 FOR APPLICANT, PART 3'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        
        row = 3
        
        # 21. Historial laboral
        ws[f'A{row}'].value = '21'
        ws[f'B{row}'].value = '職歴 Employment history'
        ws[f'B{row}'].font = self.header_font
        row += 1
        
        headers = ['勤務先名', '期間', '職種・担当業務']
        for i, h in enumerate(headers, 2):
            ws.cell(row=row, column=i).value = h
            ws.cell(row=row, column=i).font = self.small_font
            ws.cell(row=row, column=i).border = self.thin_border
        row += 1
        
        for work in data.get('work_history', []):
            ws.cell(row=row, column=2).value = work.get('company_name', '')
            ws.cell(row=row, column=3).value = work.get('period', '')
            ws.cell(row=row, column=4).value = work.get('position', '')
            for c in range(2, 5):
                ws.cell(row=row, column=c).border = self.thin_border
            row += 1
        
        row += 2
        
        # 22. Información del representante
        ws[f'A{row}'].value = '22'
        ws[f'B{row}'].value = '申請人（法定代理人）'
        ws[f'B{row}'].font = self.header_font
        row += 1
        
        ws[f'B{row}'].value = '(1) 氏名'
        rep_name = data.get('representative_name') or f"{data.get('family_name', '')} {data.get('given_name', '')}"
        ws[f'C{row}'].value = rep_name
        row += 1
        
        ws[f'B{row}'].value = '(2) 本人との関係'
        ws[f'C{row}'].value = data.get('relationship_with_applicant', '本人')
        row += 1
        
        ws[f'B{row}'].value = '(3) 住所'
        ws[f'C{row}'].value = data.get('representative_address') or data.get('address_japan', '')
        row += 1
        
        ws[f'B{row}'].value = '電話番号'
        ws[f'C{row}'].value = data.get('representative_telephone') or data.get('telephone_japan', '')
        row += 2
        
        # Firma
        ws[f'A{row}'].value = '以上の記載内容は事実と相違ありません。'
        row += 1
        ws[f'A{row}'].value = '申請人（法定代理人）の署名'
        ws[f'C{row}'].value = '___________________________'
        app_date = data.get('application_date', date.today())
        if isinstance(app_date, date):
            app_date = app_date.strftime('%Y年%m月%d日')
        ws[f'D{row}'].value = f'作成年月日: {app_date}'
    
    def _create_organization_sheet(self, wb: Workbook, data: Dict):
        """所属機関等作成用 - Información de la empresa"""
        ws = wb.create_sheet('所属機関等作成用')
        
        for i, width in enumerate([4, 15, 15, 15, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.merge_cells('A1:F1')
        ws['A1'].value = '【所属機関等作成用１】 FOR ORGANIZATION, PART 1'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        
        row = 3
        
        # 1. Nombre del trabajador extranjero
        ws[f'A{row}'].value = '1'
        ws[f'B{row}'].value = '雇用する外国人の氏名'
        foreign_name = data.get('foreign_national_name') or f"{data.get('family_name', '')} {data.get('given_name', '')}"
        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'].value = foreign_name
        row += 1
        
        # 2. Tipo de contrato
        ws[f'A{row}'].value = '2'
        ws[f'B{row}'].value = '契約の形態'
        contract = data.get('contract_type', '雇用')
        contracts = ['雇用', '委任', '請負', 'その他']
        display = '  '.join([f"☑ {c}" if c == contract else f"☐ {c}" for c in contracts])
        ws[f'C{row}'].value = display
        row += 2
        
        # 3. Información de la empresa
        ws[f'A{row}'].value = '3'
        ws[f'B{row}'].value = '所属機関等（勤務先）'
        ws[f'B{row}'].font = self.header_font
        row += 1
        
        # (1) Nombre
        ws[f'B{row}'].value = '(1) 名称'
        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'].value = data.get('employer_name', '')
        row += 1
        
        # (2) 法人番号
        ws[f'B{row}'].value = '(2) 法人番号'
        ws[f'C{row}'].value = data.get('corporation_number', '')
        row += 1
        
        # (3) 支店名
        ws[f'B{row}'].value = '(3) 支店・事業所名'
        ws[f'C{row}'].value = data.get('employer_branch', '')
        row += 1
        
        # (4) 雇用保険番号
        ws[f'B{row}'].value = '(4) 雇用保険適用事業所番号'
        ws[f'C{row}'].value = data.get('employment_insurance_number', '')
        row += 1
        
        # (5) 業種
        ws[f'B{row}'].value = '(5) 業種'
        ws[f'C{row}'].value = data.get('business_type_code', '')
        row += 1
        
        # (6) 所在地
        ws[f'B{row}'].value = '(6) 所在地'
        ws.merge_cells(f'C{row}:F{row}')
        ws[f'C{row}'].value = data.get('employer_address', '')
        row += 1
        
        # 電話番号
        ws[f'B{row}'].value = '電話番号'
        ws[f'C{row}'].value = data.get('employer_telephone', '')
        row += 1
        
        # (7) 資本金 (8) 売上
        ws[f'B{row}'].value = '(7) 資本金'
        capital = data.get('capital', '')
        ws[f'C{row}'].value = f'{capital:,}円' if capital else ''
        ws[f'D{row}'].value = '(8) 年間売上高'
        sales = data.get('annual_sales', '')
        ws[f'E{row}'].value = f'{sales:,}円' if sales else ''
        row += 1
        
        # (9) 従業員数
        ws[f'B{row}'].value = '(9) 従業員数'
        ws[f'C{row}'].value = f"{data.get('employee_count', '')}名"
        ws[f'D{row}'].value = '外国人職員数'
        ws[f'E{row}'].value = f"{data.get('foreign_employee_count', '')}名"
        row += 2
        
        # 4-8. Condiciones de empleo
        ws[f'A{row}'].value = '4'
        ws[f'B{row}'].value = '就労予定期間'
        period_type = data.get('work_period_type', '定めなし')
        ws[f'C{row}'].value = '☑ 定めなし  ☐ 定めあり' if period_type == '定めなし' else f'☐ 定めなし  ☑ 定めあり ({data.get("work_period", "")})'
        row += 1
        
        ws[f'A{row}'].value = '5'
        ws[f'B{row}'].value = '雇用開始年月日'
        start = data.get('employment_start_date', '')
        if isinstance(start, date):
            start = start.strftime('%Y年%m月%d日')
        ws[f'C{row}'].value = start
        row += 1
        
        ws[f'A{row}'].value = '6'
        ws[f'B{row}'].value = '給与・報酬（税引前）'
        salary = data.get('salary', '')
        salary_type = '月額' if data.get('salary_type') == 'monthly' else '年額'
        ws[f'C{row}'].value = f'{salary:,}円 ({salary_type})' if salary else ''
        row += 1
        
        ws[f'A{row}'].value = '7'
        ws[f'B{row}'].value = '実務経験年数'
        ws[f'C{row}'].value = f"{data.get('business_experience', '')}年"
        row += 1
        
        ws[f'A{row}'].value = '8'
        ws[f'B{row}'].value = '職務上の地位'
        ws[f'C{row}'].value = data.get('position', '')
        row += 1
        
        ws[f'A{row}'].value = '9'
        ws[f'B{row}'].value = '職種'
        ws[f'C{row}'].value = data.get('occupation_code', '')
        row += 1
        
        # 10. Detalles de actividad
        ws[f'A{row}'].value = '10'
        ws[f'B{row}'].value = '活動内容詳細'
        ws.merge_cells(f'C{row}:F{row+2}')
        ws[f'C{row}'].value = data.get('activity_details', '')
        ws[f'C{row}'].alignment = self.top_align
        row += 4
        
        # Firma de la empresa
        ws[f'A{row}'].value = '所属機関等契約先の名称・代表者氏名'
        row += 1
        ws[f'B{row}'].value = '名称:'
        ws[f'C{row}'].value = data.get('employer_name', '')
        row += 1
        ws[f'B{row}'].value = '代表者:'
        ws[f'C{row}'].value = data.get('company_representative_name', '')
        row += 1
        ws[f'B{row}'].value = '作成年月日:'
        form_date = data.get('form_creation_date', date.today())
        if isinstance(form_date, date):
            form_date = form_date.strftime('%Y年%m月%d日')
        ws[f'C{row}'].value = form_date
    
    def _add_form_row(self, ws, row: int, num: str, label_jp: str, label_en: str, value: str):
        """Agregar fila de formulario estándar"""
        ws[f'A{row}'].value = num
        ws[f'A{row}'].font = self.small_font
        ws[f'B{row}'].value = label_jp
        ws[f'B{row}'].font = self.normal_font
        ws[f'C{row}'].value = label_en
        ws[f'C{row}'].font = self.small_font
        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'].value = value
        ws[f'D{row}'].font = self.normal_font
    
    def _add_form_row_wide(self, ws, row: int, num: str, label_jp: str, label_en: str, value: str):
        """Agregar fila de formulario con valor ancho"""
        ws[f'A{row}'].value = num
        ws[f'A{row}'].font = self.small_font
        ws[f'B{row}'].value = f'{label_jp}\n{label_en}'
        ws[f'B{row}'].font = self.normal_font
        ws[f'B{row}'].alignment = self.top_align
        ws.merge_cells(f'C{row}:H{row}')
        ws[f'C{row}'].value = value
        ws[f'C{row}'].font = self.normal_font


# Función de conveniencia para generar el Excel
def generate_visa_renewal_excel(employee_data: Dict, company_data: Dict = None) -> io.BytesIO:
    """
    Genera el archivo Excel para renovación de visa
    
    Args:
        employee_data: Datos del empleado
        company_data: Datos de la empresa (opcional, se pueden incluir en employee_data)
    
    Returns:
        BytesIO con el archivo Excel
    """
    # Combinar datos
    data = {**employee_data}
    if company_data:
        data.update(company_data)
    
    generator = VisaFormExcelGenerator()
    return generator.generate_renewal_form(data)


# Ejemplo de uso
if __name__ == "__main__":
    # Datos de ejemplo
    sample_data = {
        "submission_office": "名古屋",
        "nationality": "ベトナム",
        "date_of_birth": date(1990, 5, 15),
        "family_name": "NGUYEN",
        "given_name": "VAN MINH",
        "sex": "male",
        "marital_status": "single",
        "occupation": "会社員",
        "home_town_city": "ホーチミン市",
        "address_japan": "愛知県名古屋市中区栄1-1-1",
        "telephone_japan": "052-123-4567",
        "cellular_phone": "090-1234-5678",
        "passport_number": "B12345678",
        "passport_expiration": date(2030, 1, 15),
        "current_visa_status": "技術・人文知識・国際業務",
        "current_period_of_stay": "3年",
        "current_expiration_date": date(2025, 4, 1),
        "residence_card_number": "AB12345678CD",
        "desired_visa_status": "技術・人文知識・国際業務",
        "desired_period": "5年",
        "reason_for_extension": "引き続き日本で就労するため",
        "school_name": "ハノイ工科大学",
        "school_type": "university",
        "graduation_date": date(2015, 3, 31),
        "major_field": "情報工学",
        "company_name": "株式会社UNS",
        "company_address": "愛知県春日井市○○町1-2-3",
        "company_telephone": "0568-00-0000",
        "employer_name": "株式会社UNS",
        "corporation_number": "1234567890123",
        "capital": 10000000,
        "annual_sales": 100000000,
        "employee_count": 50,
        "foreign_employee_count": 30,
        "salary": 300000,
        "salary_type": "monthly",
        "position": "正社員",
        "activity_details": "システム開発業務（Java、Python等を使用したWebアプリケーション開発）",
    }
    
    excel_file = generate_visa_renewal_excel(sample_data)
    
    # Guardar archivo
    with open("visa_renewal_form.xlsx", "wb") as f:
        f.write(excel_file.getvalue())
    
    print("✅ Excel file generated: visa_renewal_form.xlsx")
